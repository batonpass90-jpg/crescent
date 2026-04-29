# -*- coding: utf-8 -*-
"""CareOn 일괄 등록용 xlsx 양식 3종 생성"""

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, NamedStyle
)
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import os

OUT_DIR = os.path.dirname(__file__)

# ===== 공통 스타일 =====
TEAL = "0D9488"
TL = "F0FDFA"
NAVY = "0F1B2D"
GRAY = "6B7280"
LIGHT_GRAY = "F8FAFC"
RED = "DC2626"
YELLOW = "FEF3C7"
YELLOW_BORDER = "F59E0B"

def styled(ws, headers, examples, notes_rows=None):
    """공통 스타일 적용 헬퍼"""
    # 헤더
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = PatternFill("solid", fgColor=TEAL)
        c.font = Font(bold=True, color="FFFFFF", size=11, name="맑은 고딕")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(
            left=Side(style="thin", color="FFFFFF"),
            right=Side(style="thin", color="FFFFFF"),
            top=Side(style="thin", color=TEAL),
            bottom=Side(style="thin", color=TEAL),
        )

    # 예시 행 (회색 + 이탤릭)
    for i, val in enumerate(examples, 1):
        c = ws.cell(row=2, column=i, value=val)
        c.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
        c.font = Font(italic=True, color=GRAY, size=10, name="맑은 고딕")
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 24
    ws.freeze_panes = "A3"  # 헤더 + 예시 고정

def add_notes_sheet(wb, title, lines):
    """안내 시트 추가"""
    ws = wb.create_sheet(title="📋 작성 안내", index=0)
    ws.column_dimensions["A"].width = 100

    # 큰 제목
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(bold=True, size=18, color=NAVY, name="맑은 고딕")
    ws.row_dimensions[1].height = 32

    row = 3
    for line in lines:
        c = ws.cell(row=row, column=1, value=line)
        if line.startswith("【"):
            c.font = Font(bold=True, size=13, color=TEAL, name="맑은 고딕")
            c.fill = PatternFill("solid", fgColor=TL)
            ws.row_dimensions[row].height = 24
        elif line.startswith("⚠️"):
            c.font = Font(bold=True, size=11, color="92400E", name="맑은 고딕")
            c.fill = PatternFill("solid", fgColor=YELLOW)
        elif line.startswith("💡"):
            c.font = Font(size=11, color="166534", name="맑은 고딕")
            c.fill = PatternFill("solid", fgColor="DCFCE7")
        elif line == "":
            ws.row_dimensions[row].height = 8
        else:
            c.font = Font(size=11, color=NAVY, name="맑은 고딕")
        c.alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

# ============================================================
# 양식 1: 보호사 일괄 등록
# ============================================================
def create_workers_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "보호사 등록"

    headers = [
        "성함*", "연락처*", "이메일*",
        "자격증", "경력(년)", "생년월일",
        "입사일", "주소", "담당 지역",
        "급여 형태", "시급/단가/월급(원)",
        "은행 계좌"
    ]
    examples = [
        "김미영", "010-1111-2222", "kim@example.com",
        "요양보호사 1급", 5, "1985-03-15",
        "2021-04-01", "성남시 분당구 야탑동", "분당구, 수정구",
        "hourly", 12000,
        "신한 110-123-456789"
    ]
    styled(ws, headers, examples)

    # 컬럼 폭
    widths = [10, 15, 25, 15, 8, 12, 12, 25, 18, 12, 14, 22]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 데이터 검증: 급여 형태 (J열)
    dv_salary = DataValidation(
        type="list",
        formula1='"hourly,per_visit,monthly"',
        allow_blank=False
    )
    dv_salary.error = "hourly / per_visit / monthly 중 선택"
    dv_salary.errorTitle = "급여 형태 오류"
    ws.add_data_validation(dv_salary)
    dv_salary.add("J3:J100")

    # 안내 시트
    add_notes_sheet(wb, "👩‍⚕️ 보호사 일괄 등록 양식", [
        "이 시트의 다음 탭 [보호사 등록]에 보호사 정보를 입력해 주세요.",
        "1행은 헤더, 2행은 예시입니다. 3행부터 실제 데이터를 입력합니다.",
        "",
        "【필수 입력 (별표*)】",
        "  · 성함, 연락처, 이메일",
        "    이메일은 보호사가 'CareOn 보호사 앱'에 가입할 때 사용할 이메일입니다.",
        "    같은 이메일로 가입하면 자동으로 센터에 연결됩니다.",
        "",
        "【선택 입력】",
        "  · 자격증: 요양보호사 1급 / 요양보호사 2급 / 간호조무사 / 사회복지사 등",
        "  · 경력(년): 숫자만 (예: 5)",
        "  · 생년월일·입사일: YYYY-MM-DD 형식 (예: 1985-03-15)",
        "  · 담당 지역: 쉼표로 구분 (예: 분당구, 수정구)",
        "  · 급여 형태: hourly(시급) / per_visit(방문당) / monthly(월급) 중 선택",
        "  · 시급/단가/월급(원): 숫자만 (예: 12000) — 정부 수가 시급 약 12,000원",
        "",
        "【업로드 방법】",
        "  1. 파일 작성 후 저장",
        "  2. CareOn 센터장 대시보드 → 보호사 관리 → '📥 일괄 등록' 버튼",
        "  3. 이 파일 업로드 → 미리보기 확인 → '등록 확정'",
        "",
        "⚠️ 주의: 이메일이 중복되면 등록 실패합니다. 보호사 본인이 사용할 정확한 이메일을 입력해주세요.",
        "⚠️ 같은 보호사를 두 번 등록하지 마세요. 이미 등록된 보호사는 대시보드에서 직접 수정하세요.",
        "",
        "💡 보호사 등록 후 해당 이메일로 안내 문자를 자동 발송하려면 대시보드에서 '안내 발송' 버튼을 누르세요.",
        "💡 보호사 본인이 가입할 때까지 status는 'pending_signup' 상태로 표시됩니다.",
    ])

    out = os.path.join(OUT_DIR, "careon-template-workers.xlsx")
    wb.save(out)
    print(f"  saved: {out}")
    return out

# ============================================================
# 양식 2: 수급자 일괄 등록
# ============================================================
def create_patients_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "수급자 등록"

    headers = [
        "성함*", "장기요양 등급*", "생년월일",
        "연락처", "주소",
        "담당 보호사 (이메일)",
        "보호자 성함", "보호자 연락처", "보호자 관계",
        "특이사항/메모"
    ]
    examples = [
        "박종철", "2", "1948-05-20",
        "010-3333-4444", "성남시 수정구 태평동 88-12 (대평빌라 205호)",
        "kim@example.com",
        "박지영", "010-9999-8888", "배우자",
        "무릎 관절 통증, 식사 시 죽 형태 선호"
    ]
    styled(ws, headers, examples)

    widths = [10, 14, 12, 15, 32, 22, 12, 15, 12, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 등급 검증 (B열)
    dv_grade = DataValidation(
        type="list",
        formula1='"1,2,3,4,5,인지지원"',
        allow_blank=False
    )
    dv_grade.error = "1~5 또는 '인지지원' 중 선택"
    dv_grade.errorTitle = "등급 오류"
    ws.add_data_validation(dv_grade)
    dv_grade.add("B3:B100")

    # 보호자 관계 검증 (I열)
    dv_rel = DataValidation(
        type="list",
        formula1='"자녀,배우자,손주,형제자매,기타"',
        allow_blank=True
    )
    ws.add_data_validation(dv_rel)
    dv_rel.add("I3:I100")

    add_notes_sheet(wb, "👴 수급자 일괄 등록 양식", [
        "이 시트의 다음 탭 [수급자 등록]에 수급자(어르신) 정보를 입력해 주세요.",
        "1행은 헤더, 2행은 예시입니다. 3행부터 실제 데이터를 입력합니다.",
        "",
        "【필수 입력 (별표*)】",
        "  · 성함",
        "  · 장기요양 등급: 1, 2, 3, 4, 5 또는 '인지지원'",
        "",
        "【선택 입력】",
        "  · 생년월일: YYYY-MM-DD",
        "  · 연락처: 어르신 본인 휴대폰 (보호자 전용은 별도 컬럼)",
        "  · 주소: 가능한 상세 주소 (호수까지). GPS 거리 검증에 사용됩니다.",
        "",
        "【담당 보호사 매칭】",
        "  · '담당 보호사 (이메일)' 칸에 보호사 이메일을 적으면 자동 연결됩니다.",
        "  · 보호사를 먼저 등록한 다음 수급자 등록을 권장합니다.",
        "  · 미배정 시 비워두면 됩니다.",
        "",
        "【보호자 정보】",
        "  · 보호자 성함과 연락처를 적으면 자동으로 '보호자 초대' 코드가 발급됩니다.",
        "  · 코드는 SMS로 자동 전송되며, 보호자가 가입하면 자동 연결됩니다.",
        "  · 보호자 관계: 자녀 / 배우자 / 손주 / 형제자매 / 기타",
        "",
        "【업로드 방법】",
        "  1. 파일 작성 후 저장",
        "  2. CareOn 센터장 대시보드 → 수급자 관리 → '📥 일괄 등록' 버튼",
        "  3. 이 파일 업로드 → 미리보기 확인 → '등록 확정'",
        "",
        "⚠️ 같은 어르신을 두 번 등록하지 마세요. 이미 등록된 분은 대시보드에서 수정하세요.",
        "⚠️ 보호자 연락처는 보호자 본인 동의를 받은 후 입력해주세요 (개인정보).",
        "",
        "💡 등록 후 즉시 보호자에게 SMS 초대 코드가 자동 발송됩니다 (Solapi 잔액 필요).",
        "💡 첫 방문 시 보호사가 GPS 체크인하면 어르신 댁 좌표가 자동 등록됩니다.",
    ])

    out = os.path.join(OUT_DIR, "careon-template-patients.xlsx")
    wb.save(out)
    print(f"  saved: {out}")
    return out

# ============================================================
# 양식 3: 일정 일괄 등록
# ============================================================
def create_visits_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "방문 일정 등록"

    headers = [
        "방문일*", "시작 시간*", "종료 시간",
        "수급자 성함*", "보호사 이메일*",
        "서비스",
        "메모"
    ]
    examples = [
        "2026-05-02", "11:30", "13:00",
        "박종철", "kim@example.com",
        "방문요양 90분",
        "혈압 체크 필수"
    ]
    styled(ws, headers, examples)

    widths = [13, 12, 12, 12, 25, 20, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 서비스 검증 (F열)
    dv_svc = DataValidation(
        type="list",
        formula1='"방문요양 60분,방문요양 90분,방문요양 120분,방문목욕 60분,방문간호 30분"',
        allow_blank=True
    )
    ws.add_data_validation(dv_svc)
    dv_svc.add("F3:F500")

    add_notes_sheet(wb, "📅 방문 일정 일괄 등록 양식", [
        "이 시트의 다음 탭 [방문 일정 등록]에 일정을 입력해 주세요.",
        "1행은 헤더, 2행은 예시입니다. 3행부터 실제 데이터를 입력합니다.",
        "",
        "【필수 입력 (별표*)】",
        "  · 방문일: YYYY-MM-DD (예: 2026-05-02)",
        "  · 시작 시간: HH:MM (예: 11:30)",
        "  · 수급자 성함: 이미 등록된 어르신 이름과 정확히 일치해야 함",
        "  · 보호사 이메일: 등록된 보호사 이메일과 정확히 일치해야 함",
        "",
        "【선택 입력】",
        "  · 종료 시간: 미입력 시 시작 시간 + 90분 자동 계산",
        "  · 서비스: 방문요양 60/90/120분 · 방문목욕 60분 · 방문간호 30분",
        "  · 메모: 작업지시 내용 (보호사 앱에 표시됨)",
        "",
        "【반복 일정】",
        "  · 매주 월/수/금처럼 반복되는 경우 행을 복사해서 날짜만 바꿔주세요.",
        "  · 예: 박종철 어르신 매주 월수금 11:30 → 5월 한 달이면 12개 행",
        "",
        "【업로드 방법】",
        "  1. 파일 작성 후 저장",
        "  2. CareOn 센터장 대시보드 → 일정 관리 → '📥 일괄 등록' 버튼",
        "  3. 이 파일 업로드 → 미리보기 확인 → '등록 확정'",
        "",
        "⚠️ 시간 중복 체크: 같은 보호사가 같은 시간대에 두 어르신을 동시 방문하는 일정은 등록 실패합니다.",
        "⚠️ 수급자 성함 / 보호사 이메일 오타 시 등록 실패합니다. 등록된 정보와 정확히 일치해야 합니다.",
        "",
        "💡 일정 등록 후 보호사 앱에 즉시 반영되며, 작업지시는 별도로 발송 가능합니다.",
        "💡 등록된 보호사가 없으면 먼저 보호사 양식으로 보호사를 등록해주세요.",
    ])

    out = os.path.join(OUT_DIR, "careon-template-visits.xlsx")
    wb.save(out)
    print(f"  saved: {out}")
    return out

# ============================================================
if __name__ == "__main__":
    print("[CareOn] xlsx template generation starting...")
    create_workers_template()
    create_patients_template()
    create_visits_template()
    print("done.")
